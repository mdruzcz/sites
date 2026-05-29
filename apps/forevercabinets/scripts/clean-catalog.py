"""Read Cabinet Prices.xlsx, clean it, and write src/content/cabinets.json."""
import openpyxl
import json
import re
import datetime
import sys
from collections import Counter

XLSX = r"C:\Users\Matt\Downloads\Cabinet Prices.xlsx"
OUT = r"C:\Users\Matt\Documents\Claude\Projects\Website Building\sites\apps\forevercabinets\src\content\cabinets.json"


def fix_dim(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.month / v.day if v.day else None
    if isinstance(v, str):
        s = v.replace('"', "").strip()
        if s in ("/", ""):
            return None
        m = re.match(r"^(\d+)\s+(\d+)/(\d+)$", s)
        if m:
            return int(m.group(1)) + int(m.group(2)) / int(m.group(3))
        m = re.match(r"^(\d+)/(\d+)$", s)
        if m:
            return int(m.group(1)) / int(m.group(2))
        try:
            return float(s)
        except Exception:
            return None
    try:
        return float(v)
    except Exception:
        return None


def name_for(sku):
    s = str(sku).strip()
    sl = s.lower()

    if sl.startswith("lazy susan"):
        return f'{s.split()[-1]}" Lazy Susan Base'
    if sl == "stem glass holder":
        return "Under-Cabinet Stem Glass Holder"
    if sl.startswith("outside corner"):
        return "Outside Corner Moulding"
    if sl.startswith("scribe"):
        return "Scribe Moulding"
    if sl.startswith("light rail"):
        return "Light Rail Moulding"
    if sl.startswith("corbel"):
        return 'Decorative Corbel — 8" × 12"'
    if sl.startswith("touch"):
        return "Touch-Up Kit — White Shaker"
    if "48" in s and "*96" in s:
        return 'Cabinet End Panel — 48" × 96"'
    if sl.startswith("rrp"):
        return 'Refrigerator Return Panel — 24" × 96"'
    if sl.startswith("tk"):
        return 'Toe Kick — 4½" × 96"'
    if sl.startswith("dwp"):
        return "Dishwasher End Panel"

    m = re.match(r"^wf\s*(\d)\s*\*\s*(\d+)$", sl)
    if m:
        return f'Filler Strip — {m.group(1)}" × {m.group(2)}"'
    m = re.match(r"^wf(\d)(\d{2})$", sl)
    if m:
        return f'Filler Strip — {m.group(1)}" × {m.group(2)}"'
    if sl.startswith("wf"):
        return f'Filler Strip {s.replace("WF", "").strip()}'

    m = re.match(r"^B(\d+)$", s)
    if m:
        return f'{int(m.group(1))}" Base Cabinet'
    m = re.match(r"^DB(\d+)$", s)
    if m:
        return f'{int(m.group(1))}" Drawer Base'
    m = re.match(r"^SB(\d+)$", s)
    if m:
        return f'{m.group(1)}" Sink Base'
    m = re.match(r"^BBC(\d+)-(\d+)$", s)
    if m:
        return f'{m.group(1)}–{m.group(2)}" Blind Base Corner'
    m = re.match(r"^BEC(\d+)$", s)
    if m:
        return f'{m.group(1)}" Base Easy Corner'
    m = re.match(r"^BWBK(\d+)$", s)
    if m:
        return f'{m.group(1)}" Waste Basket Base'

    m = re.match(r"^WRC(\d{2})(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Glass Wine Rack Wall'
    m = re.match(r"^WP(\d{2})(\d{2})(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" × {int(m.group(3))}" Tall Pantry'
    m = re.match(r"^WDCG(\d{2})(\d{2})(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall Diagonal Corner — Glass'
    m = re.match(r"^WDC(\d+)(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall Diagonal Corner'
    m = re.match(r"^WBC(\d+)(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall Blind Corner'
    m = re.match(r"^WEC(\d+)(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall End Cabinet'
    m = re.match(r"^WGC(\d+)(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Glass Wall Cabinet'
    m = re.match(r"^WMC(\d+)(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall Microwave Cabinet'
    m = re.match(r"^W(\d{2})(\d{2})(\d{2})$", s)
    if m and int(m.group(3)) > 12:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Over-Fridge Wall ({int(m.group(3))}" depth)'
    m = re.match(r"^W(\d{2})(\d{2})$", s)
    if m:
        return f'{int(m.group(1))}" × {int(m.group(2))}" Wall Cabinet'

    return s


def slugify(s):
    s = str(s).lower().strip()
    s = (
        s.replace("½", "-half")
        .replace("¼", "-qtr")
        .replace("¾", "-3qtr")
        .replace("–", "-")
        .replace("×", "x")
    )
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def features_for(sku, name):
    feats = []
    nl = name.lower()
    if "sink" in nl:
        feats.append("sink")
    if "lazy susan" in nl:
        feats.append("lazy-susan")
    if "corner" in nl:
        feats.append("corner")
    if "waste basket" in nl:
        feats.append("waste-basket")
    if "wine rack" in nl:
        feats.append("wine-rack")
    if "glass" in nl:
        feats.append("glass-door")
    if "microwave" in nl:
        feats.append("microwave")
    if "pantry" in nl or "tall" in nl:
        feats.append("tall")
    if "drawer" in nl:
        feats.append("drawer")
    if "moulding" in nl:
        feats.append("moulding")
    if "filler" in nl:
        feats.append("filler")
    if "panel" in nl:
        feats.append("panel")
    if "toe kick" in nl:
        feats.append("toe-kick")
    if "stem" in nl:
        feats.append("stemware")
    return feats


def categorize(sku, typ_raw):
    s = str(sku).upper()
    t = (typ_raw or "").strip().lower()
    if t == "accessory":
        return "accessory"
    if s.startswith("DB"):
        return "drawer"
    if t == "base":
        return "base"
    if t == "wall":
        return "wall"
    return "accessory"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Cabinet Inventory"]
    cabinets = []
    seen = set()

    for r in range(2, ws.max_row + 1):
        typ_raw = ws.cell(row=r, column=2).value
        sku = ws.cell(row=r, column=3).value
        desc = ws.cell(row=r, column=9).value
        w = fix_dim(ws.cell(row=r, column=10).value)
        h = fix_dim(ws.cell(row=r, column=11).value)
        d = fix_dim(ws.cell(row=r, column=12).value)
        price = ws.cell(row=r, column=33).value
        if not sku or not price:
            continue
        sku_str = str(sku).strip()
        try:
            price_num = round(float(price))
        except Exception:
            continue
        cat = categorize(sku_str, typ_raw)
        name = name_for(sku_str)
        slug = slugify(sku_str)
        if slug in seen:
            slug = f"{slug}-{r}"
        seen.add(slug)
        cabinets.append(
            {
                "sku": sku_str,
                "slug": slug,
                "name": name,
                "type": cat,
                "description": str(desc).strip() if desc else None,
                "width_in": w,
                "height_in": h,
                "depth_in": d,
                "price_cad": price_num,
                "features": features_for(sku_str, name),
                "in_stock": True,
                "image_urls": [],
            }
        )

    cabinets.append(
        {
            "sku": "SAMPLE-DOOR-WS",
            "slug": "sample-door",
            "name": "White Shaker Sample Door",
            "type": "accessory",
            "description": "Order a physical sample of our White Shaker door to verify the finish matches your existing kitchen. Fully refundable on your first cabinet order.",
            "width_in": 11,
            "height_in": 15,
            "depth_in": 1,
            "price_cad": 35,
            "features": ["sample"],
            "in_stock": True,
            "image_urls": [],
        }
    )

    type_order = {"base": 0, "drawer": 1, "wall": 2, "accessory": 3}
    cabinets.sort(key=lambda c: (type_order.get(c["type"], 9), c["width_in"] or 999, c["sku"]))

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(cabinets, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(cabinets)} items to {OUT}")
    print("By type:", dict(Counter(c["type"] for c in cabinets)))
    print()
    for c in cabinets:
        print(f'  {c["type"]:9} | {c["sku"]:20} | {c["name"]}')


if __name__ == "__main__":
    main()
